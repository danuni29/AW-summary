import openpyxl as op
from copy import copy
import pandas as pd
from openpyxl.cell import cell
from openpyxl.utils.dataframe import dataframe_to_rows
from copy_cell_style import copy_row


def main():
    df = pd.read_csv('AW_summary_all_20200811.csv')
    wb = op.load_workbook('summary_template_20200812_5개년.xlsx')
    ws = wb['용수구역(통합표)']



    locations = list(df['용수구역명'].unique())


    # 이건 지우지 마삼
    # for i in range(len(locations)):
    #     copy_row(ws, i+7)

    row_idx = 7
    # avg_row_idx = 7
    # idx = 7
    #
    # five_years_temps = []
    # common_years_temps = []

    column_index = 2

    data_list = []

    for location in locations:
        ws.cell(row=row_idx, column=3).value = location   # 용수 구역명
        row_idx += 1

        cut_df = df[df['용수구역명'] == location]
        five_years_df = cut_df[(cut_df['연도'] >= 2011) & (cut_df['연도'] <= 2015)]
        common_years_df = cut_df[(cut_df['연도'] >= 1981) & (cut_df['연도'] <= 2010)]

        for idx in range(len(five_years_df.columns)):
            data = five_years_df.iloc[:,[column_index]]
            data_list.append(data)
            avg = sum(data_list)/ len(data_list)

    print(data_list)



    # print(five_years_temps)




        # five_years_temps.append(sum(five_years_df['연평균 일평균기온(℃)']) / len(five_years_df['연평균 일평균기온(℃)']))
        # common_years_temps.append(sum(common_years_df['연평균 일평균기온(℃)']) / len(common_years_df['연평균 일평균기온(℃)']))




    # for five_years_temp in five_years_temps:
    #     ws.cell(row=avg_row_idx, column=4).value = five_years_temp
    #     avg_row_idx += 1
    #
    # for common_years_temp in common_years_temps:
    #     rounded_temp = round(common_years_temp, 1)
    #     ws.cell(row=idx, column=5).value = rounded_temp
    #     idx += 1

    wb.save('summary_template_20200812_5개년.xlsx')




if __name__ == '__main__':
    main()